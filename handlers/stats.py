"""
handlers/stats.py

Полная статистика для администратора.
Многостраничная навигация + экспорт в Excel/PDF.
"""

import io
import html
from datetime import datetime

from aiogram import F, Router, Bot
from aiogram.types import (
    CallbackQuery,
    InlineKeyboardMarkup, InlineKeyboardButton,
    BufferedInputFile,
)
from aiogram.exceptions import TelegramBadRequest

from handlers.admin import ADMIN_IDS
from database.stats_models import (
    stats_users, stats_finance, stats_activity,
    stats_top_xp, stats_registrations_chart, get_full_stats,
)
from database.billing_models import get_all_plan_purchases, get_all_promo_usage
from database.models import get_user_profile


router = Router()


# =========================
# 🔧 Утилиты
# =========================
async def safe_edit(msg, text: str, markup=None):
    try:
        await msg.edit_text(text, reply_markup=markup, parse_mode='HTML')
    except TelegramBadRequest as e:
        if "message is not modified" not in str(e):
            raise


def _money(v) -> str:
    return f"{v:,.0f}₽".replace(",", " ")


def _num(v) -> str:
    return f"{v:,}".replace(",", " ")


def _bar(value: float, maximum: float, length: int = 8) -> str:
    if maximum == 0:
        return "⬜" * length
    filled = min(int(length * value / maximum), length)
    return "🟦" * filled + "⬜" * (length - filled)


# =========================
# 🎛 Клавиатуры
# =========================
def kb_stats_main() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="👥 Пользователи", callback_data="stats_users"),
            InlineKeyboardButton(text="💰 Финансы", callback_data="stats_finance"),
        ],
        [
            InlineKeyboardButton(text="📊 Активность", callback_data="stats_activity"),
            InlineKeyboardButton(text="🏆 Топ XP", callback_data="stats_top_xp"),
        ],
        [
            InlineKeyboardButton(text="📈 Динамика", callback_data="stats_chart"),
        ],
        [
            InlineKeyboardButton(text="📥 Экспорт", callback_data="stats_export_choose"),
        ],
        [
            InlineKeyboardButton(text="⬅️ Назад", callback_data="admin_main"),
        ],
    ])


def kb_export_choice() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="📊 Excel", callback_data="stats_export_excel"),
            InlineKeyboardButton(text="📄 PDF", callback_data="stats_export_pdf"),
        ],
        [
            InlineKeyboardButton(text="⬅️ Назад", callback_data="admin_statistics"),
        ],
    ])


def kb_back() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="🔄 Обновить", callback_data="stats_refresh"),
            InlineKeyboardButton(text="⬅️ К разделам", callback_data="admin_statistics"),
        ],
    ])


def kb_back_with_export() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="🔄 Обновить", callback_data="stats_refresh"),
            InlineKeyboardButton(text="📥 Excel", callback_data="stats_export_excel"),
        ],
        [InlineKeyboardButton(text="⬅️ К разделам", callback_data="admin_statistics")],
    ])


# =========================
# 🏠 Главное меню статистики
# =========================
@router.callback_query(F.data == "admin_statistics", F.from_user.id.in_(ADMIN_IDS))
async def stats_main(callback: CallbackQuery):
    await callback.answer()
    now = datetime.now().strftime("%d.%m.%Y %H:%M")
    await safe_edit(
        callback.message,
        f"📊 <b>Статистика бота</b>\n\n"
        f"Выбери раздел для просмотра.\n"
        f"<i>Обновлено: {now}</i>",
        kb_stats_main()
    )


@router.callback_query(F.data == "stats_refresh", F.from_user.id.in_(ADMIN_IDS))
async def stats_refresh(callback: CallbackQuery):
    await callback.answer("🔄 Обновлено")
    await stats_main(callback)


# =========================
# 👥 Пользователи
# =========================
@router.callback_query(F.data == "stats_users", F.from_user.id.in_(ADMIN_IDS))
async def stats_show_users(callback: CallbackQuery):
    await callback.answer()
    u = await stats_users()

    sub_pct = round(u["with_sub"] / u["total"] * 100, 1) if u["total"] else 0
    conv_bar = _bar(sub_pct, 100)

    text = (
        "👥 <b>Пользователи</b>\n"
        "━━━━━━━━━━━━━━━━━━━\n\n"

        "📌 <b>Общее</b>\n"
        f"  Всего зарегистрировано: <b>{_num(u['total'])}</b>\n"
        f"  С активной подпиской:   <b>{_num(u['with_sub'])}</b>\n"
        f"  Без подписки:           <b>{_num(u['no_sub'])}</b>\n\n"

        f"  Конверсия в подписку:\n"
        f"  {conv_bar} <b>{sub_pct}%</b>\n\n"

        "🆕 <b>Прирост</b>\n"
        f"  Сегодня:   <b>+{u['new_today']}</b>\n"
        f"  За неделю: <b>+{u['new_week']}</b>\n"
        f"  За месяц:  <b>+{u['new_month']}</b>\n"
    )
    await safe_edit(callback.message, text, kb_back())


# =========================
# 💰 Финансы
# =========================
@router.callback_query(F.data == "stats_finance", F.from_user.id.in_(ADMIN_IDS))
async def stats_show_finance(callback: CallbackQuery):
    await callback.answer()
    f = await stats_finance()

    purchases = await get_all_plan_purchases()
    promos = await get_all_promo_usage()

    # 📦 продажи по тарифам
    plan_lines = ""
    if f["plan_sales"]:
        for pname, data in sorted(f["plan_sales"].items(), key=lambda x: -x[1]["count"]):
            plan_lines += f"  • {pname}: <b>{data['count']}</b> покупок\n"
    else:
        plan_lines = "  Данных пока нет\n"

    # 🎟 промокоды (агрегировано)
    promo_lines = ""
    if f["promo_stats"]:
        for p in sorted(f["promo_stats"], key=lambda x: -x["uses"]):
            status = "✅" if p["active"] else "❌"
            promo_lines += (
                f"  {status} <code>{p['code']}</code> "
                f"−{p['discount']}%  |  использований: <b>{p['uses']}</b>\n"
            )
    else:
        promo_lines = "  Промокодов нет\n"

    # 💳 последние покупки
    purchase_users = ""
    if purchases:
        purchase_users += "\n💳 <b>Кто покупал тарифы:</b>\n"
        for p in purchases[-10:]:
            user = await get_user_profile(p["user_id"])
            username = user.get("username", "—") if user else "—"

            purchase_users += f"  • {p['plan_name']} — @{username} ({p['user_id']})\n"
    else:
        purchase_users = "\n💳 Нет покупок\n"

    # 🎟 кто использовал промо
    promo_users = ""
    if promos:
        promo_users += "\n🎟 <b>Кто использовал промокоды:</b>\n"
        for p in promos[-10:]:
            user = await get_user_profile(p["user_id"])
            username = user.get("username", "—") if user else "—"

            promo_users += f"  • {p['promo_code']} — @{username} ({p['user_id']})\n"
    else:
        promo_users = "\n🎟 Нет использований\n"

    text = (
        "💰 <b>Финансы</b>\n"
        "━━━━━━━━━━━━━━━━━━━\n\n"

        "💵 <b>Выручка</b>\n"
        f"  Сегодня:   <b>{_money(f['today_revenue'])}</b>\n"
        f"  За месяц:  <b>{_money(f['month_revenue'])}</b>\n"
        f"  Всего:     <b>{_money(f['total_revenue'])}</b>\n"
        f"  Платежей:  <b>{_num(f['payment_count'])}</b>\n\n"

        "📦 <b>Подписки</b>\n"
        f"  Активных сейчас: <b>{f['active_subs']}</b>\n\n"

        "📋 <b>Продажи по тарифам</b>\n"
        f"{plan_lines}\n"

        "🎟 <b>Промокоды</b>\n"
        f"{promo_lines}"

        f"{purchase_users}"
        f"{promo_users}"
    )

    await safe_edit(callback.message, text, kb_back_with_export())


# =========================
# 📊 Активность по разделам
# =========================
@router.callback_query(F.data == "stats_activity", F.from_user.id.in_(ADMIN_IDS))
async def stats_show_activity(callback: CallbackQuery):
    await callback.answer()
    a = await stats_activity()

    sorted_actions = sorted(a.values(), key=lambda x: -x["month"])

    lines = ""
    for item in sorted_actions:
        lines += (
            f"{item['label']}\n"
            f"  сегодня <b>{_num(item['today'])}</b>  "
            f"│ месяц <b>{_num(item['month'])}</b>  "
            f"│ всего <b>{_num(item['total'])}</b>\n\n"
        )

    if sorted_actions:
        top = sorted_actions[0]
        popular = f"🔥 Самый активный раздел: {top['label']} ({_num(top['month'])} за месяц)\n\n"
    else:
        popular = ""

    text = (
        "📊 <b>Активность по разделам</b>\n"
        "━━━━━━━━━━━━━━━━━━━\n\n"
        f"{popular}"
        f"{lines}"
    )
    await safe_edit(callback.message, text, kb_back_with_export())


# =========================
# 🏆 Топ XP
# =========================
@router.callback_query(F.data == "stats_top_xp", F.from_user.id.in_(ADMIN_IDS))
async def stats_show_top_xp(callback: CallbackQuery):
    await callback.answer()
    top = await stats_top_xp(10)

    if not top:
        await safe_edit(
            callback.message,
            "🏆 <b>Топ пользователей по XP</b>\n\n"
            "Данных пока нет — XP ещё не начислялись.",
            kb_back()
        )
        return

    medals = ["🥇", "🥈", "🥉"] + ["▪️"] * 10
    lines = ""
    for i, user in enumerate(top):
        lines += (
            f"{medals[i]} <b>{user['name']}</b>  "
            f"<code>{user['user_id']}</code>\n"
            f"   +{_num(user['xp_month'])} XP за месяц  │  "
            f"{_num(user['xp_total'])} всего\n"
            f"   <i>{user['breakdown']}</i>\n\n"
        )

    text = (
        "🏆 <b>Топ-10 пользователей за месяц</b>\n"
        "━━━━━━━━━━━━━━━━━━━\n\n"
        f"{lines}"
        "─────────────────────\n"
        "<i>Подозрительно много одного типа = возможный абуз.</i>"
    )
    await safe_edit(callback.message, text, kb_back_with_export())


# =========================
# 📈 Динамика регистраций
# =========================
@router.callback_query(F.data == "stats_chart", F.from_user.id.in_(ADMIN_IDS))
async def stats_show_chart(callback: CallbackQuery):
    await callback.answer()
    chart = await stats_registrations_chart()

    if not any(d["count"] for d in chart):
        await safe_edit(
            callback.message,
            "<b>Динамика регистраций (30 дней)</b>\n\nДанных пока нет.",
            kb_back()
        )
        return

    max_val = max(d["count"] for d in chart) or 1
    blocks = " ▁▂▃▄▅▆▇█"

    spark = ""
    for d in chart:
        idx = min(int(d["count"] / max_val * 8), 8)
        spark += blocks[idx]

    total_month = sum(d["count"] for d in chart)
    peak = max(chart, key=lambda x: x["count"])
    last7 = sum(d["count"] for d in chart[-7:])
    prev7 = sum(d["count"] for d in chart[-14:-7])
    trend = "📈" if last7 >= prev7 else "📉"
    delta = last7 - prev7

    text = (
        "📈 <b>Динамика регистраций (30 дней)</b>\n"
        "━━━━━━━━━━━━━━━━━━━\n\n"
        f"<code>{spark}</code>\n\n"
        f"  За 30 дней:   <b>{_num(total_month)}</b> новых\n"
        f"  Пик:          <b>{peak['count']}</b> ({peak['date']})\n\n"
        f"  Последние 7 дней: <b>{last7}</b>\n"
        f"  Предыдущие 7 дней: <b>{prev7}</b>\n"
        f"  Тренд: {trend} <b>{'+' if delta >= 0 else ''}{delta}</b>\n\n"
        "<i>Каждый символ = 1 день слева направо</i>"
    )
    await safe_edit(callback.message, text, kb_back_with_export())


# =========================
# 📥 Выбор формата экспорта
# =========================
@router.callback_query(F.data == "stats_export_choose", F.from_user.id.in_(ADMIN_IDS))
async def stats_export_choose(callback: CallbackQuery):
    await callback.answer()
    await safe_edit(
        callback.message,
        "📥 <b>Выбери формат экспорта:</b>",
        kb_export_choice()
    )


# =========================
# 📥 Экспорт в Excel
# =========================
@router.callback_query(F.data.in_({"stats_export_excel", "stats_export"}), F.from_user.id.in_(ADMIN_IDS))
async def stats_export_excel(callback: CallbackQuery, bot: Bot):
    await callback.answer()
    await safe_edit(callback.message, "📊 Собираю данные и формирую Excel...")

    try:
        data = await get_full_stats()
        xlsx_buf = await _build_xlsx(data)
        fname = f"stats_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

        await bot.send_document(
            chat_id=callback.from_user.id,
            document=BufferedInputFile(xlsx_buf.getvalue(), filename=fname),
            caption=(
                f"📊 <b>Статистика бота</b>\n"
                f"Сформирована: {data['generated_at']}\n\n"
                "Листы: Пользователи, Финансы, Активность, Топ XP, Динамика"
            ),
            parse_mode='HTML'
        )
        await safe_edit(
            callback.message,
            "✅ <b>Файл отправлен!</b>\n\nПроверь сообщения выше 👆",
            kb_stats_main()
        )
    except Exception as e:
        err = html.escape(str(e))[:200]
        await safe_edit(
            callback.message,
            f"❌ <b>Ошибка при формировании файла</b>\n\n<code>{err}</code>",
            kb_stats_main()
        )


# =========================
# 📄 Экспорт в PDF
# =========================
@router.callback_query(F.data == "stats_export_pdf", F.from_user.id.in_(ADMIN_IDS))
async def stats_export_pdf(callback: CallbackQuery, bot: Bot):
    await callback.answer()
    await safe_edit(callback.message, "📄 Формирую PDF...")

    try:
        data = await get_full_stats()
        pdf_buf = await _build_pdf(data)
        fname = f"stats_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"

        await bot.send_document(
            chat_id=callback.from_user.id,
            document=BufferedInputFile(pdf_buf.getvalue(), filename=fname),
            caption="📄 Статистика в PDF",
        )
        await safe_edit(callback.message, "✅ PDF отправлен!", kb_stats_main())

    except Exception as e:
        err = html.escape(str(e))[:200]
        await safe_edit(
            callback.message,
            f"❌ <b>Ошибка PDF</b>\n\n<code>{err}</code>",
            kb_stats_main()
        )


# =========================
# 🔨 Построение Excel
# =========================
async def _build_xlsx(data: dict) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference, LineChart

    wb = Workbook()
    now = data["generated_at"]

    C_HEADER  = "1F3864"
    C_SUBHEAD = "2E75B6"
    C_ACCENT  = "D6E4F7"
    C_GREEN   = "E2EFDA"
    C_YELLOW  = "FFF2CC"
    C_RED     = "FCE4D6"

    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    def header_cell(ws, row, col, text, bg=C_HEADER, fg="FFFFFF", bold=True, size=11):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(bold=bold, color=fg, name="Arial", size=size)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin
        return c

    def data_cell(ws, row, col, value, bg=None, bold=False, fmt=None, align="left"):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Arial", size=10, bold=bold)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border = thin
        if bg:
            c.fill = PatternFill("solid", start_color=bg)
        if fmt:
            c.number_format = fmt
        return c

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def title_row(ws, text, cols=6):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
        c = ws.cell(row=1, column=1, value=text)
        c.font = Font(bold=True, size=14, color="FFFFFF", name="Arial")
        c.fill = PatternFill("solid", start_color=C_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
        c2 = ws.cell(row=2, column=1, value=f"Сформировано: {now}")
        c2.font = Font(italic=True, size=9, color="666666", name="Arial")
        c2.alignment = Alignment(horizontal="center")

    # Лист 1: Пользователи
    ws1 = wb.active
    ws1.title = "Пользователи"
    title_row(ws1, "👥 Статистика пользователей", 3)
    u = data["users"]
    header_cell(ws1, 4, 1, "Метрика", bg=C_SUBHEAD)
    header_cell(ws1, 4, 2, "Значение", bg=C_SUBHEAD)
    header_cell(ws1, 4, 3, "Комментарий", bg=C_SUBHEAD)
    rows_u = [
        ("Всего пользователей", u["total"], ""),
        ("С активной подпиской", u["with_sub"],
         f"{round(u['with_sub'] / u['total'] * 100, 1) if u['total'] else 0}% конверсия"),
        ("Без подписки", u["no_sub"], ""),
        ("Новых сегодня", u["new_today"], ""),
        ("Новых за неделю", u["new_week"], ""),
        ("Новых за месяц", u["new_month"], ""),
    ]
    for i, (label, val, comment) in enumerate(rows_u, 5):
        bg = C_ACCENT if i % 2 == 0 else None
        data_cell(ws1, i, 1, label, bg=bg)
        data_cell(ws1, i, 2, val, bg=bg, bold=True, align="center")
        data_cell(ws1, i, 3, comment, bg=bg)
    set_col_widths(ws1, [32, 16, 30])

    # Лист 2: Финансы
    ws2 = wb.create_sheet("Финансы")
    title_row(ws2, "💰 Финансовая статистика", 4)
    f = data["finance"]
    header_cell(ws2, 4, 1, "Метрика", bg=C_SUBHEAD)
    header_cell(ws2, 4, 2, "Значение", bg=C_SUBHEAD)
    header_cell(ws2, 4, 3, "Период", bg=C_SUBHEAD)
    header_cell(ws2, 4, 4, "Примечание", bg=C_SUBHEAD)
    rows_f = [
        ("Выручка", f["today_revenue"], "Сегодня", "₽"),
        ("Выручка", f["month_revenue"], "Месяц", "₽"),
        ("Выручка", f["total_revenue"], "Всего", "₽"),
        ("Кол-во платежей", f["payment_count"], "Всего", "шт."),
        ("Активных подписок", f["active_subs"], "Сейчас", ""),
    ]
    for i, (label, val, period, unit) in enumerate(rows_f, 5):
        bg = C_GREEN if i % 2 == 0 else None
        data_cell(ws2, i, 1, label, bg=bg)
        data_cell(ws2, i, 2, val, bg=bg, bold=True, align="center",
                  fmt='#,##0.00' if unit == "₽" else '#,##0')
        data_cell(ws2, i, 3, period, bg=bg, align="center")
        data_cell(ws2, i, 4, unit, bg=bg, align="center")
    set_col_widths(ws2, [28, 16, 14, 20])

    # Лист 3: Активность
    ws3 = wb.create_sheet("Активность")
    title_row(ws3, "📊 Активность по разделам", 5)
    header_cell(ws3, 4, 1, "Раздел", bg=C_SUBHEAD)
    header_cell(ws3, 4, 2, "Сегодня", bg=C_SUBHEAD)
    header_cell(ws3, 4, 3, "За месяц", bg=C_SUBHEAD)
    header_cell(ws3, 4, 4, "Всего", bg=C_SUBHEAD)
    header_cell(ws3, 4, 5, "Доля месяца", bg=C_SUBHEAD)
    a = data["activity"]
    items = sorted(a.values(), key=lambda x: -x["month"])
    total_month_act = sum(x["month"] for x in items) or 1
    for i, item in enumerate(items, 5):
        bg = C_ACCENT if i % 2 == 0 else None
        share = round(item["month"] / total_month_act * 100, 1)
        data_cell(ws3, i, 1, item["label"], bg=bg)
        data_cell(ws3, i, 2, item["today"], bg=bg, align="center", bold=True)
        data_cell(ws3, i, 3, item["month"], bg=bg, align="center", bold=True)
        data_cell(ws3, i, 4, item["total"], bg=bg, align="center")
        data_cell(ws3, i, 5, f"{share}%", bg=bg, align="center")
    set_col_widths(ws3, [26, 12, 14, 14, 16])

    # Лист 4: Топ XP
    ws4 = wb.create_sheet("Топ XP")
    title_row(ws4, "🏆 Топ пользователей по XP за месяц", 6)
    header_cell(ws4, 4, 1, "#", bg=C_SUBHEAD)
    header_cell(ws4, 4, 2, "Имя", bg=C_SUBHEAD)
    header_cell(ws4, 4, 3, "User ID", bg=C_SUBHEAD)
    header_cell(ws4, 4, 4, "XP месяц", bg=C_SUBHEAD)
    header_cell(ws4, 4, 5, "XP всего", bg=C_SUBHEAD)
    header_cell(ws4, 4, 6, "Разбивка действий", bg=C_SUBHEAD)
    medals_bg = [C_YELLOW, "D9D9D9", "F4CCCC"] + [None] * 20
    for i, user in enumerate(data["top_xp"], 5):
        rank = i - 4
        bg = medals_bg[rank - 1]
        data_cell(ws4, i, 1, rank, bg=bg, bold=(rank <= 3), align="center")
        data_cell(ws4, i, 2, user["name"], bg=bg, bold=(rank <= 3))
        data_cell(ws4, i, 3, str(user["user_id"]), bg=bg, align="center")
        data_cell(ws4, i, 4, user["xp_month"], bg=bg, bold=True, align="center", fmt='#,##0')
        data_cell(ws4, i, 5, user["xp_total"], bg=bg, align="center", fmt='#,##0')
        data_cell(ws4, i, 6, user["breakdown"], bg=bg)
    if not data["top_xp"]:
        data_cell(ws4, 5, 1, "Нет данных")
    set_col_widths(ws4, [6, 22, 14, 12, 12, 40])

    # Лист 5: Динамика
    ws5 = wb.create_sheet("Динамика")
    title_row(ws5, "📈 Динамика регистраций (30 дней)", 2)
    header_cell(ws5, 4, 1, "Дата", bg=C_SUBHEAD)
    header_cell(ws5, 4, 2, "Новых", bg=C_SUBHEAD)
    for i, item in enumerate(data["reg_chart"], 5):
        bg = C_GREEN if item["count"] > 0 else None
        data_cell(ws5, i, 1, item["date"], bg=bg)
        data_cell(ws5, i, 2, item["count"], bg=bg, bold=item["count"] > 0, align="center")
    total_row = len(data["reg_chart"]) + 5
    data_cell(ws5, total_row, 1, "ИТОГО за 30 дней", bold=True, bg=C_YELLOW)
    ws5.cell(row=total_row, column=2).value = f'=SUM(B5:B{total_row - 1})'
    ws5.cell(row=total_row, column=2).font = Font(bold=True, name="Arial", size=10)
    ws5.cell(row=total_row, column=2).fill = PatternFill("solid", start_color=C_YELLOW)
    ws5.cell(row=total_row, column=2).alignment = Alignment(horizontal="center")
    ws5.cell(row=total_row, column=2).border = thin
    set_col_widths(ws5, [14, 10])

    # График динамики
    chart = LineChart()
    chart.title = "Регистрации по дням"
    chart.y_axis.title = "Пользователи"
    chart.x_axis.title = "Дата"
    data_ref = Reference(ws5, min_col=2, min_row=4, max_row=total_row - 1)
    cats_ref = Reference(ws5, min_col=1, min_row=5, max_row=total_row - 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 10
    chart.width = 22
    ws5.add_chart(chart, "D4")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# =========================
# 📄 Построение PDF
# =========================
def _generate_insights(data: dict) -> list:
    insights = []
    u = data["users"]
    f = data["finance"]

    if u["total"]:
        conv = u["with_sub"] / u["total"] * 100
        insights.append(f"Конверсия: {conv:.1f}%")

    avg = f["month_revenue"] / 30 if f["month_revenue"] else 0
    today = f["today_revenue"]
    if avg > 0:
        delta_pct = (today - avg) / avg * 100
        if delta_pct > 0:
            insights.append(f"Выручка выросла на {delta_pct:.1f}%")
        else:
            insights.append(f"Выручка упала на {abs(delta_pct):.1f}%")

    last7 = sum(x["count"] for x in data["reg_chart"][-7:])
    prev7 = sum(x["count"] for x in data["reg_chart"][-14:-7])
    if prev7 > 0:
        delta = (last7 - prev7) / prev7 * 100
        if delta > 0:
            insights.append(f"Регистрации выросли на {delta:.1f}%")
        else:
            insights.append(f"Регистрации упали на {abs(delta):.1f}%")

    if data["top_xp"]:
        top = data["top_xp"][0]
        if top["xp_month"] > 5000:
            insights.append(f"Аномально высокий XP у {top['name']}")

    return insights


async def _build_pdf(data: dict) -> io.BytesIO:
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.pagesizes import A4

        # Пробуем зарегистрировать шрифт, если нет — используем Helvetica
        try:
            pdfmetrics.registerFont(TTFont("DejaVu", "media/fonts/DejaVuSans.ttf"))
            font_name = "DejaVu"
        except Exception:
            font_name = "Helvetica"

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Statistika bota", styles["Title"]))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph(f"Sformirowano: {data['generated_at']}", styles["Normal"]))
        elements.append(Spacer(1, 20))

        # Авто-анализ
        insights = _generate_insights(data)
        if insights:
            elements.append(Paragraph("Avto-analiz", styles["Heading2"]))
            for ins in insights:
                elements.append(Paragraph(ins, styles["Normal"]))
            elements.append(Spacer(1, 20))

        # Пользователи
        u = data["users"]
        elements.append(Paragraph("Polzovateli", styles["Heading2"]))
        table = Table([
            ["Metrika", "Znachenie"],
            ["Vsego", u["total"]],
            ["S podpiskoy", u["with_sub"]],
            ["Bez podpiski", u["no_sub"]],
            ["Segodnya", u["new_today"]],
            ["Nedelya", u["new_week"]],
            ["Mesyac", u["new_month"]],
        ])
        table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))

        # Финансы
        f = data["finance"]
        elements.append(Paragraph("Finansy", styles["Heading2"]))
        table = Table([
            ["Metrika", "Znachenie"],
            ["Segodnya", f["today_revenue"]],
            ["Mesyac", f["month_revenue"]],
            ["Vsego", f["total_revenue"]],
            ["Platezhi", f["payment_count"]],
            ["Aktivnye podpiski", f["active_subs"]],
        ])
        table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ]))
        elements.append(table)

        doc.build(elements)
        buf.seek(0)
        return buf

    except ImportError:
        # Если reportlab не установлен — возвращаем простой текстовый PDF через fpdf2
        # или просто текстовый файл с расширением pdf
        buf = io.BytesIO()
        text = f"Stats report {data['generated_at']}\n\n"
        text += f"Users total: {data['users']['total']}\n"
        text += f"With sub: {data['users']['with_sub']}\n"
        text += f"Revenue total: {data['finance']['total_revenue']}\n"
        buf.write(text.encode('utf-8'))
        buf.seek(0)
        return buf